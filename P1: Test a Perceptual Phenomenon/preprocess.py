import openpyxl
from openpyxl import Workbook

book = openpyxl.load_workbook('Book2.xlsx')

sheet = book.active
wb = Workbook()
ws = wb.active

# a = sheet['A4']
# print a.value

row_range = sheet['A1':'A307']
# for row in row_range:
# 	for cell in row:
# 		print (cell.value)

i=0
for row in row_range:
	i = i+1
	for cell in row:
		a,b,c,d = cell.value.split(',')
		#print a,b,c,d
		ws.cell(row=i, column=1).value = a
		ws.cell(row=i, column=2).value = b
		ws.cell(row=i, column=3).value = c
		ws.cell(row=i, column=4).value = d

wb.save('prep.csv')

# print sheet.cell(row=45, column=1).value

